import requests
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import os
import re
import getpass





print("=+=+=+=+=+=+   Getting started   =+=+=+=+=+=+\n\n")

perseverance = ["Foundries", "Steel industries", "Casting Industries"]
endurance = {"Gujarat": ['Ahmedabad', 'Amreli', 'Anand', 'Aravalli', 'Banaskantha', 'Bharuch', 'Bhavnagar', 'Botad', 'Chhota Udepur', 'Dahod','Dang', 'Dev Bhoomi Dwarka', 'Gandhinagar', 'Gir Somnath', 'Jamnagar', 'Junagadh', 'Kachchh', 'Kheda', 'Mahisagar', 'Mehsana', 'Morbi', 'Kheda', 'Mahisagar', 'Mehsana', 'Morbi', 'Narmada', 'Navsari', 'Panchmahal', 'Patan', 'Porbandar', 'Rajkot', 'Sabarkantha', 'Surat', 'Surendranagar', 'Tapi', 'Vadodara', 'Valsad']}

username = getpass.getuser()

path = 'C:\\Users\\' + username + '\\Desktop'
os.chdir(path)

if not os.path.isdir("Industry"):
    os.mkdir("Industry")
os.chdir('Industry')


cumTotalCount = 0
cumStart = time.time()

#perseverance = ["Foundries"]
#endurance = {"Tamil Nadu": ["Coimbatore"]}

for query in perseverance:

    searchQuery = query
    searchQuery = searchQuery.replace(' ', '+').lower()

    if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query):
        os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query)


    queryStart = time.time()
    print("\n\n\nCURRENT QUERY => " + query + "\n\n\n")
    queryLeavel = 0

    for state in endurance:

        searchState = state
        searchState = searchState.replace(' ', '+').lower()

        if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query + '\\' + state):
            os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query + '\\' + state)

        stateStart = time.time()
        print("\n\nCURRENT STATE => " + state + "\n\n")
        stateLeavel = 0
        districts = endurance[state]

        for district in districts:

            searchDistrict = district
            searchDistrict = searchDistrict.replace(' ', '+').lower()

            if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query + '\\' + state + '\\' + district):
                os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\Industry\\' + query + '\\' + state + '\\' + district)

            districtStart = time.time()
            nameList = []
            districtLevel = 0
            count = 0

            print("\n====================================\nEXTRACTING " + query + " IN " + str(district).upper() + "\n")

            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet.title = query + ' in ' + district + '(v2.0.1)'

            sheet.cell(1, 1).value = "Sr No."
            sheet.cell(1, 2).value = "Name of " + query
            sheet.cell(1, 3).value = "Mobile No."
            sheet.cell(1, 4).value = "Ph No."
            sheet.cell(1, 5).value = "Address"
            sheet.cell(1, 6).value = "Type"
            sheet.cell(1, 7).value = "Stars"
            sheet.cell(1, 8).value = "No. Votes"
            sheet.cell(1, 9).value = "Website"
            sheet.cell(1, 10).value = "WebLink1"
            sheet.cell(1, 11).value = "WebLink2"
            sheet.cell(1, 12).value = "WebLink3"

            link = 'https://www.google.com/search?q=' + searchQuery + '+in+' + searchDistrict + '+' + searchState
            a = requests.get(link)
            ht = open('read.html', 'w')
            ht.write(a.text)
            ht.close()
            soup = bs(a.text, "html.parser")

            try:
                mainLink = \
                soup.find('div', {'id': 'main'}).find('div', {'class': 'ZINbbc xpd O9g5cc uUPGi'}).findAll('div', recursive=False)[
                    -1].find('a')['href']
                browser = webdriver.Chrome()

                mainLink = 'https://www.google.com' + mainLink
                #print("\n-----" + mainLink + "-----\n")

                browser.maximize_window()
                browser.get(mainLink)
            except:
                print("\n********Error extracting,", query, state, districts, "*********\n")

            # (Update 25 Sep 2021, suddenly stopped working, so I added https:/www.google.com before the mainLink
            # at this time, main link outpouts, /search?ie=UTF-8&tbs=lf:1,lf_ui:2&q=foundries+in+coimbatore+tamil+nadu&rlst=f&rflfq=1&num=10&sa=X&ved=2ahUKEwi376jupZjzAhUcxjgGHVw3ANgQjGp6BAgFEAw

            j = 1  #j is the page counter

            while True:

                print("PAGE: " + str(j))
                print('----------\n')
                WebDriverWait(browser, 10).until(EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="rlfl__tls rl_tls"]/*')))
                time.sleep(3)

                try:
                    seq = browser.find_element_by_xpath('//div[@class="rlfl__tls rl_tls"]') #(25 sep 2021, stopped working so used xpath instead of selector)
                except:
                    print("***************** Failed to extract page " + str(j) + " in district: " + district, '***********')
                    continue

                childs = seq.find_elements_by_xpath("*")
                NuOfChilds = len(childs)
                i = 1 #i counts the number of elements on a particular page

                for child in childs:

                    if j == 1 and i == NuOfChilds-1:
                        break
                    if i == NuOfChilds:
                        break


                    indchild = child.find_element_by_class_name('cXedhc')
                    indchild = indchild.find_elements_by_xpath("*")

                    name = indchild[0].text


                    other = indchild[1]
                    info = other.find_elements_by_tag_name('div')

                    try:
                        stars = info[0].find_element_by_tag_name('span').text
                        votes = info[0].find_elements_by_tag_name('span')[-1].text
                        type = info[0].text.split(' · ')[-1].strip()
                        #ERROR WITH THE ABOVE LINE, CODE DEPLOYEDV2 INITIALISED
                        #list index out of range error

                    except:
                        stars = 'No reviews'
                        votes = 'No reviews'
                        type = info[0].text.split(' · ')[-1].strip()

                    try:
                        address = info[1].text.split(' · ')[-1]
                    except:
                        address = "-"
                    try:
                        phno = info[2].text.split(' · ')[-1]
                        if len(phno.split()) == 2:
                            mno = phno
                            mno = mno[1:len(mno) - 1]
                            phno = '-'
                        else:
                            mno = '-'
                    except:
                        phno = "-"
                        mno = '-'


                    #---------Start of the onclick method-----------
                    # Any comment in the onclick applies to the special click for last element as well
                    try:
                        WebDriverWait(browser, 10).until(
                            EC.invisibility_of_element_located((By.XPATH, '//div[@class="QU77pf"]')))
                    except:
                        print("Error with the invisibility of the cross mark")


                    #Waiting for the cross mark button to be gone from the screen
                    #Remove it if required, I think it will help us to mitigate the issue of
                    #selenium.common.exceptions.ElementClickInterceptedException: Message: element click intercepted: Element <div class="cXedhc">...</div> is not clickable at point (159, 780). Other element would receive the click: <div class="wYWDAd"></div>

                    while True:
                        try:
                            actions = ActionChains(browser)
                            actions.move_to_element(child).perform()
                            #foundary = child.find_element_by_xpath('.//div[@class="cXedhc"]')
                            child.click()
                            break
                        except:
                            continue

                    try:
                        close = WebDriverWait(browser, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                    except:
                        print("Error while getting the crossmark")
                    #Finding the close button prehand togive enough time to other elements

                    websitePresent = 1
                    webLinkPresent = 1

                    try:
                        #WebDriverWait(browser, 3).until(
                            #EC.visibility_of_all_elements_located((By.XPATH, '//a[@class="ab_button CL9Uqc"]')))

                        website = WebDriverWait(browser, 1.5).until(
                            EC.visibility_of_element_located((By.LINK_TEXT, 'Website')))

                        #Replace the above line with the rest and give website to the first line if required, second line not added due
                            #major requirement
                            #even changer website by website[0], if done

                        #if website[0].text != 'Website':
                            #websitePresent = 0
                    except:
                        websitePresent = 0

                    try:
                        webLink = WebDriverWait(browser, 1.5).until(
                            EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="QjJaxe Nx3I"]')))
                    except:
                        webLinkPresent = 0

                    if websitePresent:
                        websiteValue = website.get_attribute('href')
                        #print(websiteValue)
                    else:
                        websiteValue = '-'

                    webLinkList = []
                    if webLinkPresent:
                        webLinkIndex = 0
                        for webLinkF in webLink:
                            webLinkIndex+=1
                            webLinkList.append(webLinkF.find_element_by_xpath('.//a[1]').get_attribute('href'))
                    else:
                        webLinkList = ['-', '-', '-']

                    try:
                        #close = WebDriverWait(browser, 3).until(
                            #EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                        #close button was moved to the top

                        close.click()
                    except:
                        pass
                        #print("issue with close button")

                    #----------------End of the part of the clicking--------------------


                    if (name in nameList):
                        i+=1
                        continue

                    nameList.append(name)
                    #print(name)
                    #print(websiteValue)
                    #print(webLinkList)
                    print("Extracted number: " + str(count))
                    count += 1

                    sheet.cell(count + 1, 1).value = count
                    sheet.cell(count + 1, 2).value = name
                    sheet.cell(count + 1, 3).value = mno
                    sheet.cell(count + 1, 4).value = phno
                    sheet.cell(count + 1, 5).value = address
                    sheet.cell(count + 1, 6).value = type
                    sheet.cell(count + 1, 7).value = stars
                    sheet.cell(count + 1, 8).value = votes
                    sheet.cell(count + 1, 9).value = websiteValue

                    for counter in range(webLinkIndex):
                        sheet.cell(count + 1, 10+counter).value = webLinkList[counter]

                    if webLinkIndex == 2:
                        sheet.cell(count + 1, 12).value = '-'
                    if webLinkIndex == 1:
                        sheet.cell(count + 1, 11).value = '-'
                        sheet.cell(count + 1, 12).value = '-'
                    if webLinkIndex == 0:
                        sheet.cell(count + 1, 10).value = '-'
                        sheet.cell(count + 1, 11).value = '-'
                        sheet.cell(count + 1, 12).value = '-'

                    i += 1


                try:
                    pleasewait = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Next')))
                    # print("Error while waiting2")
                    next = browser.find_element_by_link_text('Next')
                    # print("Error while waiting3")
                    next.click()
                    # wait = WebDriverWait(browser, 60).until((EC.invisibility_of_element((By.LINK_TEXT, 'Next'))))
                    # print("Error while waiting4")


                except Exception as e:
                    # All of a sudden there is no need to add a last child, so I commented out the following code of adding a last child
                    # update of 26sep 2021, so I added a try except block insted of just commenting out the code
                    try:
                        last = childs[-1]
                        last = last.find_element_by_class_name('cXedhc')
                        indchild = last.find_elements_by_xpath("*")


                        #ADDING THE LAST CHILD
                        name = indchild[0].text

                        other = indchild[1]
                        info = other.find_elements_by_tag_name('div')

                        try:
                            stars = info[0].find_element_by_tag_name('span').text
                            votes = info[0].find_elements_by_tag_name('span')[-1].text
                            type = info[0].text.split(' · ')[-1].strip()

                        except Exception as exp:
                            stars = 'No reviews'
                            votes = 'No reviews'
                            type = info[0].text.split(' · ')[-1].strip()


                        try:
                            address = info[1].text.split(' · ')[-1]
                        except:
                            address = "-"

                        try:
                            phno = info[2].text.split('·')[-1]
                            if (len(phno.split()) == 2):
                                mno = phno
                                mno = mno[1:len(mno) - 1]
                                phno = '-'
                            else:
                                mno = '-'
                        except:
                            phno = "-"
                            mno = '-'

                        if name in nameList:
                            wb.save('.//' + query + '//' + state + '//' + district + '//' + district + '_v(2.0.1).xlsx')
                            print('\n-----', query, 'added in', district, state, '-----\n')
                            wb.close()
                            break


                        #special click for the last element

                        try:
                            WebDriverWait(browser, 10).until(
                                EC.invisibility_of_element_located((By.XPATH, '//div[@class="QU77pf"]')))
                        except:
                            print("Error with the inivisibility of the cross mark for the cumLastElement")

                        lastchildren = childs[-1]

                        while True:
                            try:
                                actions = ActionChains(browser)
                                actions.move_to_element(lastchildren).perform()
                                # foundary = child.find_element_by_xpath('.//div[@class="cXedhc"]')
                                lastchildren.click()
                                break
                            except:
                                continue


                        #foundary = lastchildren.find_element_by_xpath('.//div[@class="cXedhc"]')
                        #foundary.click()

                        try:
                            close = WebDriverWait(browser, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                        except:
                            print("Error with the clickability of the cross mark")

                        websitePresent = 1
                        webLinkPresent = 1

                        try:
                            #website = WebDriverWait(browser, 3).until(
                                #EC.visibility_of_all_elements_located((By.XPATH, '//a[@class="ab_button CL9Uqc"]')))

                            website = WebDriverWait(browser, 1.5).until(
                                EC.visibility_of_element_located((By.LINK_TEXT, 'Website')))

                            #if website[0].text != 'Website':
                                #websitePresent = 0
                        except:
                            websitePresent = 0

                        try:
                            webLink = WebDriverWait(browser, 1.5).until(
                                EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="QjJaxe Nx3I"]')))
                        except:
                            webLinkPresent = 0

                        if websitePresent:
                            websiteValue = website.get_attribute('href')
                            #print(websiteValue)

                        webLinkList = []
                        if webLinkPresent:
                            webLinkIndex = 0
                            for webLinkF in webLink:
                                webLinkIndex += 1
                                webLinkList.append(webLinkF.find_element_by_xpath('.//a[1]').get_attribute('href'))

                        else:
                            webLinkList = ['-', '-', '-']

                        try:
                            close.click()
                        except:
                            pass
                            #print("issue with close button")

                        #end of special click for the last element



                        nameList.append(name)
                        #print(websiteValue)
                        #print(webLinkList)
                        #print(votes)
                        #print(type)
                        #print(address)
                        #print(phno)
                        print("Extracted number: " + str(count))
                        count += 1

                        sheet.cell(count + 1, 1).value = count
                        sheet.cell(count + 1, 2).value = name
                        sheet.cell(count + 1, 3).value = mno
                        sheet.cell(count + 1, 4).value = phno
                        sheet.cell(count + 1, 5).value = address
                        sheet.cell(count + 1, 6).value = type
                        sheet.cell(count + 1, 7).value = stars
                        sheet.cell(count + 1, 8).value = votes
                        sheet.cell(count + 1, 9).value = websiteValue

                        for counter in range(webLinkIndex):
                            sheet.cell(count + 1, 10 + counter).value = webLinkList[counter]

                        if webLinkIndex == 2:
                            sheet.cell(count + 1, 12).value = '-'
                        if webLinkIndex == 1:
                            sheet.cell(count + 1, 11).value = '-'
                            sheet.cell(count + 1, 12).value = '-'
                        if webLinkIndex == 0:
                            sheet.cell(count + 1, 10).value = '-'
                            sheet.cell(count + 1, 11).value = '-'
                            sheet.cell(count + 1, 12).value = '-'

                        i+=1

                        # ADDED THE LAST CHILD
                    except Exception as e:
                        #print(e)
                        pass

                    wb.save('.//' + query + '//' + state + '//' + district + '//' + district + '_v(2.0.1).xlsx')
                    print('\n-----', query, 'added in', district, state, '-----\n')
                    wb.close()
                    break

                j += 1



            print("\n====================================\nEXTRACTED " + query + " IN " + str(district).upper() + "\n")
            districtLevel += count
            browser.close()

            stateLeavel += districtLevel
            print("\n===================\n")
            print("Total extracted foundries: " + str(districtLevel - 1))
            print("Time taken: " + str(time.time() - districtStart) + " seconds")
            print("===================\n")

        queryLeavel += stateLeavel

        print("\n\n===============================================\n\n")
        print(state, "=> completed")
        print("Total extracted foundries: " + str(stateLeavel - 1))
        print("Time taken: " + str(time.time() - stateStart) + " seconds")
        print("===============================================\n\n")

    cumTotalCount += queryLeavel

    print("\n\n\n===============================================")
    print(query, "=> completed")
    print("Total extracted foundries: " + str(queryLeavel - 1))
    print("Time taken: " + str(time.time() - queryStart) + " seconds")
    print("===============================================\n\n\n")


print("\n\n===============================================\n")
print("Job Done")
print("Total extracted foundries: " + str(cumTotalCount - 1))
print("Time taken: " + str(time.time() - cumStart) + " seconds")
print("\n===============================================")


'''
Error A 


IN case you get the error related to the version of the chrome driver, for example
This version of ChromeDriver only supports Chrome version 87
Current browser version is 93.0.4577.82 with binary path C:\Program Files (x86)\Google\Chrome\Application\chrome.exe

179

I solved these kinds of problems using the webdrive manager.

You can automatically use the correct chromedriver by using the webdrive-manager. Install the webdrive-manager:

pip install webdriver-manager
Then use the driver in python as follows

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(ChromeDriverManager().install())
This answer is taken from https://stackoverflow.com/a/52878725/10741023

This is a possible answer as well: https://stackoverflow.com/a/49795348

'''

'''req = requests.get(mainLink)
soup = bs(req.text, 'html.parser')

mainweb = open('mainweb.html', 'w')
mainweb.write(req.text)

map = soup.find('div', {'id':'main'}).findAll('div', recursive=False)[2].find('a')['href']

req_map = requests.get(map)
mainmap = open('map.html', 'w')
mainmap.write(req_map.text)
print(req_map.text)'''

'''
body = soup.find('body')
print(body)

y = body.find('div', {'class':'main'})
print(y)
#main = body.find('div', {'class':'main', 'id':',main'}, recursive=True)
#print(main)
#next = soup.find('td', {'class':'d6cvqb'})
#print(next)'''

'''
WEBDRIVER wait condition

WebDriverWait wait = new WebDriverWait(webDriver, timeoutInSeconds);
wait.until(ExpectedConditions.visibilityOfElementLocated(By.id<locator>));

'''

'''

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

browser = webdriver.Firefox()
browser.get("url")
delay = 3 # seconds
try:
    myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'IdOfMyElement')))
    print "Page is ready!"
except TimeoutException:
    print "Loading took too much time!"

'''





'''
CODE: RED
Selenium scroll until the element appears

from selenium.webdriver.common.action_chains import ActionChains

element = driver.find_element_by_id("my-id")

actions = ActionChains(driver)
actions.move_to_element(element).perform()

'''














