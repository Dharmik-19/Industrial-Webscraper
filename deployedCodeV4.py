import requests
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options  #V3, for an error called , chrome not reachable
import time
from datetime import date
from datetime import datetime
import openpyxl
import os
import re
import getpass

#The latest updates to the code are marked as V3




print("=+=+=+=+=+=+   Getting started   =+=+=+=+=+=+\n\n")



#perseverance = ["Foundries", "Steel industries", "Casting Industries"]
#endurance = {"Gujarat": ['Ahmedabad', 'Amreli', 'Anand', 'Aravalli', 'Banaskantha', 'Bharuch', 'Bhavnagar', 'Botad', 'Chhota Udepur', 'Dahod', 'Dang', 'DevBhoomi Dwarka', 'Gandhinagar', 'Gir Somnath', 'Jamnagar', 'Junagadh', 'Kachchh', 'Kheda', 'Mahisagar', 'Mehsana', 'Morbi', 'Narmada', 'Navsari', 'Panchmahal', 'Patan', 'Porbandar', 'Rajkot', 'Sabarkantha', 'Surat', 'Surendranagar', 'Tapi', 'Vadodara', 'Valsad']}
#Above ones for dad

#Below ones for me
#perseverance = ["Mining Industries", "Petrochemical Firms", "Food Processing Firms", "Oil and Natural Gas Industries", "Pharmaceutical Firms", "Fertilizer Firm", "Chemical Firms", "Battery Manufacturers", "Paint Manufacturers"]
perseverance = ["Battery Manufacturers", "Paint Manufacturers"]
#endurance = {"Gujarat": ['Anand', 'Bharuch', 'Dahod', 'Dang', 'DevBhoomi Dwarka', 'Kachchh', 'Surendranagar', 'Tapi', 'Valsad', 'Panchmahal']}
endurance = {"Andhra Pradesh": ["Andhra Pradesh"], "Arunachal Pradesh":["Arunachal Pradesh"], "Assam":["Assam"], "Bihar":["Bihar"], "Chhattisgarh":["Chhattisgarh"], "Goa":["Goa"], "Gujarat":["Gujarat"], "Haryana":["Haryana"], "Himachal Pradesh":["Himachal Pradesh"], "Jharkhand":["Jharkhand"], "Karnataka":["Karnataka"], "Kerala":["Kerala"], "Madhya Pradesh":["Madhya Pradesh"], "Maharashtra":["Maharashtra"], "Manipur":["Manipur"], "Meghalaya":["Meghalaya"], "Mizoram":["Mizoram"], "Nagaland":["Nagaland"], "Odisha":["Odisha"], "Punjab":["Punjab"], "Rajasthan":["Rajasthan"], "Sikkim":["Sikkim"], "Tamil Nadu":["Tamil Nadu"], "Telangana":["Telangana"], "Tripura":["Tripura"], "Uttar Pradesh":["Uttar Pradesh"], "Uttarakhand":["Uttarakhand"], "West Bengal": ["West Bengal"], "Andaman and Nicobar Islands": ["Andaman and Nicobar Islands"], "Chandigarh": ["Chandigarh"], "Dadra and Nagar Haveli and Daman and Diu": ["Dadra and Nagar Haveli and Daman and Diu"], "Delhi": ["Delhi"], "Jammu and Kashmir": ["Jammu and Kashmir"], "Ladakh": ["Ladakh"], "Lakshadweep": ["Lakshadweep"], "Puducherry": ["Puducherry"]}

username = getpass.getuser()

path = 'C:\\Users\\' + username + '\\Desktop'
os.chdir(path)

if not os.path.isdir("chemIndustryV5"):
    os.mkdir("chemIndustryV5")
os.chdir('chemIndustryV5')

log = open('log.txt', 'w')
log.write(str(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\tCode execution initialised\n"))
log.write("===============================\n\n\n")
log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tInitializing path and creating folders\n\n')
log.flush()

cumTotalCount = 0
cumStart = time.time()

#perseverance = ["Foundries"]
#endurance = {"Tamil Nadu": ["Coimbatore"]}

for query in perseverance:
    queryStart = time.time()
    log.write('\n==============================================================\n')
    log.write('==============================================================\n')
    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tCurrent Query:' + ' ' + query + ' ' + '\n')
    log.write('==============================================================\n')
    log.write('==============================================================\n\n')
    log.flush()

    searchQuery = query
    searchQuery = searchQuery.replace(' ', '+').lower()

    if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query):
        os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query)


    queryStart = time.time()
    print("\n\n\nCURRENT QUERY => " + query + "\n\n\n")
    queryLevel = 0

    for state in endurance:


        searchState = state
        searchState = searchState.replace(' ', '+').lower()

        if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query + '\\' + state):
            os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query + '\\' + state)

        stateStart = time.time()
        print("\n\nCURRENT STATE => " + state + "\n\n")
        log.write('\n=======================================\n')
        log.write('=======================================\n')
        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}) Current State:'.format(query) + ' ' + state + ' ' + '\n')
        log.write('=======================================\n')
        log.write('=======================================\n\n')
        log.flush()
        stateLevel = 0
        districts = endurance[state]

        nameList = []

        for district in districts:
            districtStart = time.time()
            log.write('\n=======================================\n')
            log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}) Current District:'.format(query, state) + ' ' + district + ' ' + '\n')
            log.write('=======================================\n\n')
            log.flush()

            searchDistrict = district
            searchDistrict = searchDistrict.replace(' ', '+').lower()

            if not os.path.isdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query + '\\' + state + '\\' + district):
                os.mkdir('C:\\Users\\' + username + '\\Desktop' + '\\chemIndustryV5\\' + query + '\\' + state + '\\' + district)



            districtLevel = 0
            count = 0

            print("\n====================================\nEXTRACTING " + query + " IN " + str(district).upper() + "\n")

            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet.title = query + ' in ' + district + '(v2.0.1)'

            sheet.cell(1, 1).value = "Sr No."
            sheet.cell(1, 2).value = "Name of " + query
            sheet.cell(1, 3).value = "Mobile No."
            sheet.cell(1, 4).value = "Ph No."
            sheet.cell(1, 5).value = "Address"
            sheet.cell(1, 6).value = "Type"
            sheet.cell(1, 7).value = "Stars"
            sheet.cell(1, 8).value = "No. Votes"
            sheet.cell(1, 9).value = "Website"
            sheet.cell(1, 10).value = "WebLink1"
            sheet.cell(1, 11).value = "WebLink2"
            sheet.cell(1, 12).value = "WebLink3"

            link = 'https://www.google.com/search?q=' + searchQuery + '+in+' + searchDistrict + '+' + searchState
            a = requests.get(link)
            # ht = open('read.html', 'w')
            # ht.write(a.text)
            # ht.close()
            #soup = bs(a.text, "html.parser")
            # htmlFile = open('html.html', 'wb')
            # htmlFile.write(soup.prettify('UTF-8'))
            # htmlFile.close()

            log.write(str(date.today()) + ' ' + str(
                datetime.now().time()) + ':' + ' ' + "\t({}>{}>{})Initialising headless chrome browser\n".format(query,
                                                                                                                  state,
                                                                                                                  district))
            log.flush()

            options = Options()#V3
            #options.headless = True
            options.add_argument("--no-sandbox")#V3
            options.add_argument("--disable-setuid-sandbox")#V3
            #V4 (Changed chrome_options to options)

            browser = webdriver.Chrome(options=options)#V4 -> Changed (chrome_options = chrome_options to the present value)
            browser.maximize_window()
            browser.get(link)
            #C:\Python39\Scripts Here  is the chrome driver files, the chromedriver exe which would be over there would be getting executed

            success_in_link_of_district = 0
            try:
                mainLink = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@class="ndElDd"]/g-more-link/a[@class="tiS4rf Q2MMlc"]')))
                success_in_link_of_district = 1
            except:
                pass

            if success_in_link_of_district:
                mainLink.click()
            else:
                print("\n********Error extracting Link,", query, state, district, "*********\n")
                log.write('\n****************************\n')
                log.write(str(date.today()) + ' ' + str(
                    datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) Error extracting link and chrome browser,".format(
                    query, state,
                    district) + ' ' + query + ' ' + state + ' ' + district + "\n")
                log.write('****************************\n\n')
                log.flush()
                browser.close()
                continue
                #finalUpdates


            # for trying in range(10):
            #
            #     #mainLink = soup.find('html').find('a', {'class': 'tiS4rf Q2MMlc'})['href']
            #     #mainLink = soup.find('div', {'id': 'main'})
            #
            #     try:
            #         #mainLink = soup.find('div', {'id': 'main'}).find('div', {'class': 'ZINbbc xpd O9g5cc uUPGi'}).findAll('div', recursive=False)[
            #             #-1].find('a')['href']
            #         #updated the main link because I think it is inaccurate
            #         mainLink = soup.find('div', {'class': 'ndElDd'}).find('a', {'class': 'tiS4rf Q2MMlc'})['href']
            #
            #         mainLink = 'https://www.google.com' + mainLink
            #         print(mainLink)       #dummy
            #         success_in_link_of_element = 1
            #         break
            #
            #     except:
            #         time.sleep(1)
            #         continue


            # if success_in_link_of_element:
            #     browser = webdriver.Chrome()
            #     browser.maximize_window()
            #     browser.get(mainLink)
            # else:
            #     print("\n********Error extracting Link,", query, state, district, "*********\n")
            #     log.write('\n****************************\n')
            #     log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) Error extracting link and chrome browser,".format(query, state,
            #                                                                                      district) + ' ' + query + ' ' + state + ' ' + district + " *********\n\n")
            #     log.write('****************************\n\n')
            #     log.flush()
            #     continue



            # (Update 25 Sep 2021, suddenly stopped working, so I added https:/www.google.com before the mainLink
            # at this time, main link outpouts, /search?ie=UTF-8&tbs=lf:1,lf_ui:2&q=foundries+in+coimbatore+tamil+nadu&rlst=f&rflfq=1&num=10&sa=X&ved=2ahUKEwi376jupZjzAhUcxjgGHVw3ANgQjGp6BAgFEAw

            j = 1  #j is the page counter

            #break_out_from_district = 0

            while True:

                print("PAGE: " + str(j))
                print('----------\n')
                log.write('\n' + str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) PAGE: ".format(query, state, district) + str(j) + ' ' + '\n')
                log.write('-----------------------\n')
                log.flush()

                try:
                    WebDriverWait(browser, 120).until(EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="rlfl__tls rl_tls"]/*')))
                except:
                    pass

                time.sleep(3)

                try:
                    WebDriverWait(browser, 120).until(EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="rlfl__tls rl_tls"]/*')))
                    seq = browser.find_element_by_xpath('//div[@class="rlfl__tls rl_tls"]') #(25 sep 2021, stopped working so used xpath instead of selector)
                    childs = seq.find_elements_by_xpath("*")
                    #V3 - 60 to 120
                except:
                    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}>{}) Failed to extract page'.format(query, state, district) + ' ' + str(j) + ' ' + '\n')
                    log.flush()
                    print("***************** Failed to extract page " + str(j) + " in district: " + district, '***********')

                    #break_out_from_district = 1
                    print("BROKE OUT FROM THE PAGINATION")
                    wb.save('.//' + query + '//' + state + '//' + district + '//' + district + '_v(2.0.1).xlsx')
                    break
                    #final updates

                #if break_out_from_district:
                    #break   #TO break out from the for loop for districts and shift to the next one


                NuOfChilds = len(childs)
                i = 1 #i counts the number of elements on a particular page


                for child in childs:

                    if j == 1 and i == NuOfChilds-1:
                        break
                    if i == NuOfChilds:
                        break

                    try:
                        indchild = child.find_element_by_class_name('cXedhc')
                        indchild = indchild.find_elements_by_xpath("*")

                        name = indchild[0].text


                        other = indchild[1]
                        info = other.find_elements_by_tag_name('div')
                    except:
                        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) **** Failed to extract element:".format(query, state, district) + ' ' + str(count) + ' ' + '****\n')
                        log.flush()

                    try:
                        stars = info[0].find_element_by_tag_name('span').text
                        votes = info[0].find_elements_by_tag_name('span')[-1].text
                        type = info[0].text.split(' · ')[-1].strip()

                    except:
                        stars = 'No reviews'
                        votes = 'No reviews'
                        try:
                            type = info[0].text.split(' · ')[-1].strip()
                        except:
                            type = '-'

                    try:
                        address = info[1].text.split(' · ')[-1]
                    except:
                        address = "-"

                    try:
                        phno = info[2].text.split(' · ')[-1]
                        if len(phno.split()) == 2:
                            mno = phno
                            mno = mno[1:len(mno)]
                            phno = '-'
                        else:
                            mno = '-'
                    except:
                        phno = "-"
                        mno = '-'


                    #---------Start of the onclick method-----------
                    # Any comment in the onclick applies to the special click for last element as well
                    try:
                        WebDriverWait(browser, 10).until(
                            EC.invisibility_of_element_located((By.XPATH, '//div[@class="QU77pf"]')))
                    except:
                        print("Error with the invisibility of the cross mark")


                    #Waiting for the cross mark button to be gone from the screen
                    #Remove it if required, I think it will help us to mitigate the issue of
                    #selenium.common.exceptions.ElementClickInterceptedException: Message: element click intercepted: Element <div class="cXedhc">...</div> is not clickable at point (159, 780). Other element would receive the click: <div class="wYWDAd"></div>

                    for temp1 in range(10):
                        try:
                            actions = ActionChains(browser)
                            actions.move_to_element(child).perform()
                            #foundary = child.find_element_by_xpath('.//div[@class="cXedhc"]')
                            child.click()
                            break
                        except:
                            time.sleep(1)
                            continue

                    try:
                        close = WebDriverWait(browser, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                    except:
                        print("Error while getting the crossmark")
                    #Finding the close button prehand togive enough time to other elements

                    websitePresent = 1
                    webLinkPresent = 1

                    try:
                        #WebDriverWait(browser, 3).until(
                            #EC.visibility_of_all_elements_located((By.XPATH, '//a[@class="ab_button CL9Uqc"]')))

                        website = WebDriverWait(browser, 1.5).until(
                            EC.visibility_of_element_located((By.LINK_TEXT, 'Website')))
                        websiteLink = website.get_attribute('href')

                        #Replace the above line with the rest and give website to the first line if required, second line not added due
                            #major requirement
                            #even changer website by website[0], if done

                        #if website[0].text != 'Website':
                            #websitePresent = 0
                    except:
                        websitePresent = 0

                    try:
                        webLink = WebDriverWait(browser, 1.5).until(
                            EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="QjJaxe Nx3I"]')))

                        webLinkList = []
                        webLinkIndex = 0
                        for webLinkF in webLink:
                            webLinkIndex += 1
                            webLinkList.append(webLinkF.find_element_by_xpath('.//a[1]').get_attribute('href'))

                        if webLinkIndex == 0:
                            webLinkPresent = 0

                    except:
                        webLinkPresent = 0



                    if websitePresent:
                        websiteValue = websiteLink
                        #print(websiteValue)
                    else:
                        websiteValue = '-'


                    if webLinkPresent:
                        pass

                    else:
                        webLinkList = ['-', '-', '-']

                    try:
                        #close = WebDriverWait(browser, 3).until(
                            #EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                        #close button was moved to the top

                        close.click()
                    except:
                        pass
                        #print("issue with close button")

                    #----------------End of the part of the clicking--------------------


                    if (name in nameList):
                        i+=1
                        continue

                    nameList.append(name)
                    #print(name)
                    #print(websiteValue)
                    #print(webLinkList)
                    print("Extracted number: " + str(count))
                    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) Extracted number:".format(query, state, district) + ' ' + str(count) + ' ' + '\n')
                    log.flush()
                    count += 1

                    sheet.cell(count + 1, 1).value = count
                    sheet.cell(count + 1, 2).value = name
                    sheet.cell(count + 1, 3).value = mno
                    sheet.cell(count + 1, 4).value = phno
                    sheet.cell(count + 1, 5).value = address
                    sheet.cell(count + 1, 6).value = type
                    sheet.cell(count + 1, 7).value = stars
                    sheet.cell(count + 1, 8).value = votes
                    sheet.cell(count + 1, 9).value = websiteValue
                    #print(mno, phno)

                    for counter in range(webLinkIndex):
                        sheet.cell(count + 1, 10+counter).value = webLinkList[counter]

                    if webLinkIndex == 2:
                        sheet.cell(count + 1, 12).value = '-'
                    if webLinkIndex == 1:
                        sheet.cell(count + 1, 11).value = '-'
                        sheet.cell(count + 1, 12).value = '-'
                    if webLinkIndex == 0:
                        sheet.cell(count + 1, 10).value = '-'
                        sheet.cell(count + 1, 11).value = '-'
                        sheet.cell(count + 1, 12).value = '-'

                    i += 1


                try:
                    #pleasewait = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Next')))
                    #pleasewait = WebDriverWait(browser, 10).unitl(EC.element_to_be_clickable((By.XPATH, '//table[@class="AaVjTc"]/tbody/tr/td[@class="d6cvqb"][2]')))
                    print("FInals stand")


                    #Resume work from here, pleasewait variable is not able to find the exactl location of hte element under studcy, preferablly use the find element by link text for Next

                    pleasewait = WebDriverWait(browser, 10).until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="rl_ist0"]/div/div[2]/div/table/tbody/tr/td[last()]/a')))
                    #Got a problem due to the 12 in the bottom line, as few website do not have pagination upto index 12. So we will use the above line; amazing
                    #We are at V4
                    #pleasewait = WebDriverWait(browser, 10).until(EC.element_to_be_clickable(
                        #(By.XPATH, '//*[@id="rl_ist0"]/div/div[2]/div/table/tbody/tr/td[12]/a')))
                    #Fuck, all of a suffed this is able to find the element, check hte below line, xpath is the same in the below and the above line
                    #NO FUCKING IDEA


                    #pleasewait = WebDriverWait(browser, 10).unitl(EC.element_to_be_clickable((By.XPATH, '//*[@id="rl_ist0"]/div/div[2]/div/table/tbody/tr/td[12]')))
                    #used the aboslute xpath instead of manually selecting it using 2 lines above}
                    # print("Error while waiting2")
                    #next = browser.find_element_by_link_text('Next')
                    next = pleasewait
                    #Thee above two lines are updations to the ones above em
                    #This could help to mitigate overshootting of pagenation issue.
                    #finalUpdates

                    # print("Error while waiting3")
                    next.click()
                    # wait = WebDriverWait(browser, 60).until((EC.invisibility_of_element((By.LINK_TEXT, 'Next'))))
                    # print("Error while waiting4")


                except Exception as e:
                    # All of a sudden there is no need to add a last child, so I commented out the following code of adding a last child
                    # update of 26sep 2021, so I added a try except block insted of just commenting out the code
                    try:
                        last = childs[-1]
                        last = last.find_element_by_class_name('cXedhc')
                        indchild = last.find_elements_by_xpath("*")


                        #ADDING THE LAST CHILD
                        name = indchild[0].text

                        other = indchild[1]
                        info = other.find_elements_by_tag_name('div')

                        try:
                            stars = info[0].find_element_by_tag_name('span').text
                            votes = info[0].find_elements_by_tag_name('span')[-1].text
                            type = info[0].text.split(' · ')[-1].strip()

                        except Exception as exp:
                            stars = 'No reviews'
                            votes = 'No reviews'
                            try:
                                type = info[0].text.split(' · ')[-1].strip()
                            except:
                                type = '-'

                        try:
                            address = info[1].text.split(' · ')[-1]
                        except:
                            address = "-"

                        try:
                            phno = info[2].text.split('·')[-1]
                            if (len(phno.split()) == 2):
                                mno = phno
                                mno = mno[1:len(mno)]
                                phno = '-'
                            else:
                                mno = '-'
                        except:
                            phno = "-"
                            mno = '-'

                        if name in nameList:
                            wb.save('.//' + query + '//' + state + '//' + district + '//' + district + '_v(2.0.1).xlsx')
                            print('\n-----', query, 'added in', district, state, '-----\n')
                            log.write('\n--------------\n')
                            log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}>{})'.format(query, state, district) + ' ' + query + ' ' + 'added in' + ' ' + district + ' ' + state + ' ' + '\n')
                            log.write('--------------\n\n')
                            log.flush()
                            wb.close()
                            break


                        #special click for the last element

                        try:
                            WebDriverWait(browser, 10).until(
                                EC.invisibility_of_element_located((By.XPATH, '//div[@class="QU77pf"]')))
                        except:
                            print("Error with the inivisibility of the cross mark for the cumLastElement")

                        lastchildren = childs[-1]

                        for temp2 in range(10):
                            try:
                                actions = ActionChains(browser)
                                actions.move_to_element(lastchildren).perform()
                                # foundary = child.find_element_by_xpath('.//div[@class="cXedhc"]')
                                lastchildren.click()
                                break
                            except:
                                time.sleep(1)
                                continue


                        #foundary = lastchildren.find_element_by_xpath('.//div[@class="cXedhc"]')
                        #foundary.click()

                        try:
                            close = WebDriverWait(browser, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '//div[@class="QU77pf"]')))
                        except:
                            print("Error with the clickability of the cross mark")

                        websitePresent = 1
                        webLinkPresent = 1

                        try:
                            #website = WebDriverWait(browser, 3).until(
                                #EC.visibility_of_all_elements_located((By.XPATH, '//a[@class="ab_button CL9Uqc"]')))

                            website = WebDriverWait(browser, 1.5).until(
                                EC.visibility_of_element_located((By.LINK_TEXT, 'Website')))
                            websiteLink = website.get_attribute('href')

                            #if website[0].text != 'Website':
                                #websitePresent = 0
                        except:
                            websitePresent = 0

                        try:
                            webLink = WebDriverWait(browser, 1.5).until(
                                EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="QjJaxe Nx3I"]')))

                            webLinkList = []
                            webLinkIndex = 0
                            for webLinkF in webLink:
                                webLinkIndex += 1
                                webLinkList.append(webLinkF.find_element_by_xpath('.//a[1]').get_attribute('href'))

                            if webLinkIndex == 0:
                                webLinkPresent = 0

                        except:
                            webLinkPresent = 0

                        if websitePresent:
                            websiteValue = websiteLink
                            #print(websiteValue)
                        else:
                            websiteValue = '-'

                        if webLinkPresent:
                            pass

                        else:
                            webLinkList = ['-', '-', '-']

                        try:
                            close.click()
                        except:
                            pass
                            #print("issue with close button")

                        #end of special click for the last element



                        nameList.append(name)
                        #print(websiteValue)
                        #print(webLinkList)
                        #print(votes)
                        #print(type)
                        #print(address)
                        #print(phno)
                        print("Extracted number: " + str(count))
                        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + "\t({}>{}>{}) Extracted number:".format(query, state, district) + ' ' + str(count) + ' ' + '\n')
                        log.flush()
                        count += 1

                        sheet.cell(count + 1, 1).value = count
                        sheet.cell(count + 1, 2).value = name
                        sheet.cell(count + 1, 3).value = mno
                        sheet.cell(count + 1, 4).value = phno
                        sheet.cell(count + 1, 5).value = address
                        sheet.cell(count + 1, 6).value = type
                        sheet.cell(count + 1, 7).value = stars
                        sheet.cell(count + 1, 8).value = votes
                        sheet.cell(count + 1, 9).value = websiteValue
                        #print(mno, phno)


                        for counter in range(webLinkIndex):
                            sheet.cell(count + 1, 10 + counter).value = webLinkList[counter]

                        if webLinkIndex == 2:
                            sheet.cell(count + 1, 12).value = '-'
                        if webLinkIndex == 1:
                            sheet.cell(count + 1, 11).value = '-'
                            sheet.cell(count + 1, 12).value = '-'
                        if webLinkIndex == 0:
                            sheet.cell(count + 1, 10).value = '-'
                            sheet.cell(count + 1, 11).value = '-'
                            sheet.cell(count + 1, 12).value = '-'

                        i+=1

                        # ADDED THE LAST CHILD
                    except Exception as e:
                        #print(e)
                        pass

                    wb.save('.//' + query + '//' + state + '//' + district + '//' + district + '_v(2.0.1).xlsx')
                    log.write('\n--------------\n')
                    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}>{})'.format(query, state, district) + ' ' + query + ' ' + 'added in' + ' ' + district + ' ' + state + ' ' + '\n')
                    log.write('--------------\n\n')
                    log.flush()
                    wb.close()
                    break

                j += 1

            districtLevel += count

            print("\n====================================\nEXTRACTED " + query + " IN " + str(district).upper() + "\n")
            log.write('\n=======================================\n')
            log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}) EXTRACTED '.format(query, state) + query + ' IN ' + str(district).upper() + ' ' + '\n')
            log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}) Total extracted {}:'.format(query, state, query) + ' ' + str(districtLevel - 1) + ' ' + '\n')
            log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}>{}) Time taken:'.format(query, state) + ' ' + str(time.time()-districtStart) + ' ' + 'seconds\n')
            log.write('=======================================\n\n')
            log.flush()
            wb.close()
            browser.close()

            stateLevel += districtLevel
            print("\n===================\n")
            print("Total extracted foundries: " + str(districtLevel - 1))
            print("Time taken: " + str(time.time() - districtStart) + " seconds")
            print("===================\n")

        queryLevel += stateLevel


        log.write('\n=======================================\n')
        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}) EXTRACTED '.format(query) + query + ' IN ' + str(state).upper() + ' ' + '\n')
        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}) Total extracted {}:'.format(query, query) + ' ' + str(stateLevel - 1) + ' ' + '\n')
        log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\t({}) Time taken:'.format(query) + ' ' + str(time.time() - stateStart) + ' ' + 'seconds\n')
        log.write('=======================================\n\n')
        log.flush()

        print("\n\n===============================================\n\n")
        print(state, "=> completed")
        print("Total extracted foundries: " + str(stateLevel - 1))
        print("Time taken: " + str(time.time() - stateStart) + " seconds")
        print("===============================================\n\n")

    log.write('\n==============================================================\n')
    log.write('==============================================================\n')
    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tEXTRACTED ' + query.upper() + ' ' + '\n')
    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tTotal extracted {}:'.format(query) + ' ' + str(queryLevel - 1) + ' ' + '\n')
    log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tTime taken:' + ' ' + str(time.time() - queryStart) + ' ' + 'seconds\n')
    log.write('==============================================================\n')
    log.write('==============================================================\n\n')
    log.flush()
    cumTotalCount += queryLevel

    print("\n\n\n===============================================")
    print(query, "=> completed")
    print("Total extracted {}: ".format(query) + str(queryLevel - 1))
    print("Time taken: " + str(time.time() - queryStart) + " seconds")
    print("===============================================\n\n\n")




print("\n\n===============================================\n")
print("Job Done")
print("Total extracted industries: " + str(cumTotalCount - 1))
print("Time taken: " + str(time.time() - cumStart) + " seconds")
print("\n===============================================")

log.write("\n\n========================================\n")
log.write("+=+=+=+=+=+   Code Executed Successfully   =+=+=+=+=+=+\n")
log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tTotal extracted industries:' + ' ' + str(cumTotalCount - 1) + ' ' + '\n')
log.write(str(date.today()) + ' ' + str(datetime.now().time()) + ':' + ' ' + '\tTime taken:' + ' ' + str(time.time() - cumStart) + ' ' + 'seconds\n')
log.write("========================================\n")
log.close()
'''
Error A 


IN case example
This version of ChromeDriver only supports Chrome version  you get the error related to the version of the chrome driver, for87
Current browser version is 93.0.4577.82 with binary path C:\Program Files (x86)\Google\Chrome\Application\chrome.exe

179

I solved these kinds of problems using the webdrive manager.

You can automatically use the correct chromedriver by using the webdrive-manager. Install the webdrive-manager:

pip install webdriver-manager
Then use the driver in python as follows

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(ChromeDriverManager().install())
This answer is taken from https://stackoverflow.com/a/52878725/10741023

This is a possible answer as well: https://stackoverflow.com/a/49795348

'''

'''req = requests.get(mainLink)
soup = bs(req.text, 'html.parser')

mainweb = open('mainweb.html', 'w')
mainweb.write(req.text)

map = soup.find('div', {'id':'main'}).findAll('div', recursive=False)[2].find('a')['href']

req_map = requests.get(map)
mainmap = open('map.html', 'w')
mainmap.write(req_map.text)
print(req_map.text)'''

'''
body = soup.find('body')
print(body)

y = body.find('div', {'class':'main'})
print(y)
#main = body.find('div', {'class':'main', 'id':',main'}, recursive=True)
#print(main)
#next = soup.find('td', {'class':'d6cvqb'})
#print(next)'''

'''
WEBDRIVER wait condition

WebDriverWait wait = new WebDriverWait(webDriver, timeoutInSeconds);
wait.until(ExpectedConditions.visibilityOfElementLocated(By.id<locator>));

'''

'''

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

browser = webdriver.Firefox()
browser.get("url")
delay = 3 # seconds
try:
    myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'IdOfMyElement')))
    print "Page is ready!"
except TimeoutException:
    print "Loading took too much time!"

'''





'''
CODE: RED
Selenium scroll until the element appears

from selenium.webdriver.common.action_chains import ActionChains

element = driver.find_element_by_id("my-id")

actions = ActionChains(driver)
actions.move_to_element(element).perform()

'''














